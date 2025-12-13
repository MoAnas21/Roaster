from config import config, inputs, constraints, no_days, employees, shift_colours, start_date, end_date
from generate_roaster import simulate_roaster
from feasibility_checker import check_feasibility
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Validate inputs before starting
def validate_inputs(config, inputs, constraints):
    """Validate that all required inputs are present and correctly formatted."""
    errors = []
    
    # Check required config keys
    required_config = ["no_employees", "no_shifts", "work_pattern", "forbidden_constraints"]
    for key in required_config:
        if key not in config:
            errors.append(f"Missing config key: {key}")
    
    # Check required input keys
    required_inputs = ["shift_day", "work_pattern", "previous_day", "quality_count"]
    for key in required_inputs:
        if key not in inputs:
            errors.append(f"Missing input key: {key}")
    
    # Validate array lengths
    if "no_employees" in config:
        n_emp = config["no_employees"]
        for key in ["shift_day", "work_pattern", "previous_day", "quality_count"]:
            if key in inputs and len(inputs[key]) != n_emp:
                errors.append(f"Input '{key}' length ({len(inputs[key])}) doesn't match number of employees ({n_emp})")
        
        # Validate employee_leaves if present
        if "employee_leaves" in inputs:
            if len(inputs["employee_leaves"]) != n_emp:
                errors.append(f"Input 'employee_leaves' length ({len(inputs['employee_leaves'])}) doesn't match number of employees ({n_emp})")
            else:
                # Validate each leave set
                for i, leave_set in enumerate(inputs["employee_leaves"]):
                    if not isinstance(leave_set, set):
                        errors.append(f"Employee {i}: employee_leaves[{i}] must be a set")
                    else:
                        # Check that all leave day indices are valid (non-negative integers)
                        for day_idx in leave_set:
                            if not isinstance(day_idx, int) or day_idx < 0:
                                errors.append(f"Employee {i}: employee_leaves[{i}] contains invalid day index: {day_idx}")
        
        # Validate shift_preferences if present
        if "shift_preferences" in inputs:
            if len(inputs["shift_preferences"]) != n_emp:
                errors.append(f"Input 'shift_preferences' length ({len(inputs['shift_preferences'])}) doesn't match number of employees ({n_emp})")
            else:
                # Validate each preference set
                valid_shift_ids = set(range(1, config.get("no_shifts", 0) + 1))
                for i, pref_set in enumerate(inputs["shift_preferences"]):
                    if not isinstance(pref_set, set):
                        errors.append(f"Employee {i}: shift_preferences[{i}] must be a set")
                    else:
                        # Check that all shift IDs are valid
                        for shift_id in pref_set:
                            if not isinstance(shift_id, int):
                                errors.append(f"Employee {i}: shift_preferences[{i}] contains non-integer: {shift_id}")
                            elif shift_id not in valid_shift_ids and shift_id != 0:
                                errors.append(f"Employee {i}: shift_preferences[{i}] contains invalid shift ID: {shift_id}. Valid IDs: {valid_shift_ids}")
        
        # Validate quality_count structure
        if "quality_count" in inputs:
            for i, qc in enumerate(inputs["quality_count"]):
                if "no_shifts" in config and len(qc) != config["no_shifts"]:
                    errors.append(f"Employee {i} quality_count length ({len(qc)}) doesn't match number of shifts ({config['no_shifts']})")
    
    # Validate constraints
    if "min_count" in constraints and "max_count" in constraints:
        for shift_id in range(1, config.get("no_shifts", 0) + 1):
            if shift_id in constraints["min_count"] and shift_id in constraints["max_count"]:
                if constraints["min_count"][shift_id] > constraints["max_count"][shift_id]:
                    errors.append(f"Shift {shift_id}: min_count ({constraints['min_count'][shift_id]}) > max_count ({constraints['max_count'][shift_id]})")
    
    if errors:
        raise ValueError("Input validation failed:\n" + "\n".join(f"  - {e}" for e in errors))
    
    return True

# Initialize schedule
inputs["schedule"] = []

# Validate inputs
print("Validating inputs...")
try:
    validate_inputs(config, inputs, constraints)
    print("✓ Input validation passed")
except ValueError as e:
    print(f"✗ Validation error: {e}")
    raise

# Check feasibility
print("\nChecking feasibility...")
is_feasible, feasibility_messages = check_feasibility(config, inputs, constraints, no_days)

if feasibility_messages:
    print("\nFeasibility Check Results:")
    for msg in feasibility_messages:
        if msg.startswith("INFEASIBLE"):
            print(f"  ✗ {msg}")
        else:
            print(f"  ⚠ {msg}")
    
    if not is_feasible:
        print("\n✗ Problem is INFEASIBLE. Please adjust constraints, leaves, or work patterns.")
        raise ValueError("Problem is infeasible - see feasibility check results above")
    else:
        print("\n⚠ Warnings detected but problem may still be solvable.")

print("\nStarting simulation...")
print(f"Days to schedule: {no_days}")
print(f"Employees: {config['no_employees']}")
print(f"Shifts: {config['no_shifts']}")
print(f"Date range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

final_solutions, final_quality_count = simulate_roaster(0, no_days, config, inputs, constraints)

if final_solutions is not None:
    # Combine previous day (if exists) with generated schedule
    if "previous_day" in inputs and inputs["previous_day"]:
        added_final_solution = [inputs["previous_day"]] + final_solutions
    else:
        added_final_solution = final_solutions
    data = {}
    for i, day in enumerate(added_final_solution):
        # column_name = f"Day {i + 1}"
        current_date_obj = start_date + pd.Timedelta(days=i)
        column_name = current_date_obj.strftime("%Y-%m-%d")
        data[column_name] = [f"Shift {val}" if val != 0 else "Off" for val in day]

    data["Employee name"] = [employees[j]["name"] for j in range(0, len(added_final_solution[0]))]
    data["Employee id"] = [employees[j]["employee_id"] for j in range(0, len(added_final_solution[0]))]
    data["Work Pattern"] = [employees[j]["preferred_work_pattern"] for j in range(0, len(added_final_solution[0]))]

    data_columns = [(start_date + pd.Timedelta(days=i)).strftime("%Y-%m-%d") for i in range(0, len(added_final_solution))]
    columns = ["Employee id", "Employee name", "Work Pattern"] + data_columns
    df = pd.DataFrame(data, columns=columns)
    df.to_csv("roaster.csv", index=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Roaster"

    # Move "Employee name" to the last column
    columns = [col for col in data_columns] + ["Employee id", "Employee name", "Work Pattern"]

    for col_num, column_title in enumerate(columns, start=1):
        ws.cell(row=1, column=col_num, value=column_title)

    for row_num, employee_data in enumerate(zip(*[data[col] for col in columns]), start=2):
        for col_num, cell_value in enumerate(employee_data, start=1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    employee_name_color = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    value_colors = {}
    color_index = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.column >= ws.max_column-3:  # Last column (Employee name)
                cell.fill = employee_name_color
            else:
                value_colors[cell.value] = PatternFill(start_color=shift_colours[cell.value],
                    end_color=shift_colours[cell.value],
                    fill_type="solid")
                cell.fill = value_colors[cell.value]

    # Move "Employee name" to the first column
    columns = ["Employee id", "Employee name", "Work Pattern"] + data_columns

    for col_num, column_title in enumerate(columns, start=1):
        ws.cell(row=1, column=col_num, value=column_title)

    for row_num, employee_data in enumerate(zip(*[data[col] for col in columns]), start=2):
        for col_num, cell_value in enumerate(employee_data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            if col_num <= 3:  # First column (Employee name)
                cell.fill = employee_name_color
            else:
                cell.fill = value_colors.get(cell_value, PatternFill(fill_type=None))
    wb.save("roaster.xlsx")
    print(f"\n✓ Roaster generated successfully!")
    print(f"  - CSV saved: roaster.csv")
    print(f"  - Excel saved: roaster.xlsx")
else:
    print("\n✗ Failed to generate roaster schedule.")
    print("  Possible reasons:")
    print("  - Constraints too strict (min/max employee counts)")
    print("  - Work patterns incompatible with date range")
    print("  - Insufficient employees for shift requirements")
    print("  - Try increasing 'threshold' in config or relaxing constraints")